"""
Workbook class tests
"""

import unittest
from pathlib import Path

import openpyxl

from workbook import Workbook


class WorkbookTests(unittest.TestCase):
    """
    Test function classes ability to find functions and to handle incorrect inputs
    """

    def setUp(self):
        self.test_workbook = Workbook()
        self.test_workbook.work_book = openpyxl.load_workbook(Path('.') / 'tests' /
                                                              'test' / 'test.xlsx')
        self.test_workbook.work_sheet = self.test_workbook.work_book.active
        self.test_workbook.wb_name = 'test'

    def test_choose_existing_worklist(self):
        """
        Test existing worklist can be chosen
        """
        expected = self.test_workbook.get_worklist().title
        actual = openpyxl.load_workbook(Path('.') / 'tests' / 'test' / 'test.xlsx').active.title
        self.assertEqual(expected, actual)

    def test_save_file(self):
        """
        Test file saving
        """
        expected = True
        actual = False
        self.test_workbook.save()
        saved_file = Path('.') / 'work' / 'test_result.xlsx'
        if saved_file.exists():
            actual = True
        self.assertEqual(expected, actual)
