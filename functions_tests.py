"""
Test for main functions
"""
import unittest
from pathlib import Path

import openpyxl
from functions import Functions, CountingFunctions, DifficultFunctions, MovingFunctions, \
    DeleteFunction, CompareFunction, SearchingFunction


class DefineFunctionsTest(unittest.TestCase):
    """
    Test function classes ability to find functions and to handle incorrect inputs
    """

    def setUp(self):
        self.test_sheet = openpyxl.load_workbook(Path('.') / 'tests' / 'test' / 'test.xlsx').active
        self.functions_types = {'counting functions':
                                ['add', 'sub', 'multi', 'div', 'mean'],
                                'difficult functions': ['rounded', 'expo', 'log'],
                                'moving functions': ['move', 'copy'],
                                'delete function': ['delete'],
                                'compare function': ['compare'],
                                'searching function': ['find']}
        self.bad_inputs = ['string', 1, 0.5, [], {}, (), None, [None]]

    def test_function_class_defines_functions(self):
        """
        Functions class can define functions
        """
        for function_type, function_names in self.functions_types.items():
            expected = function_type
            for function_name in function_names:
                func = Functions(function_name, self.test_sheet)
                func.define_function_type()
                actual = func.get_called_type()
                self.assertEqual(expected, actual)

    def test_function_class_handles_bad_inputs(self):
        """
        Functions class can handle bad inputs
        """
        expected = -1
        for bad_input in self.bad_inputs:
            func = Functions(bad_input, self.test_sheet)
            actual = func.define_function_type()
            self.assertEqual(expected, actual)


class CellPatternCheckTest(unittest.TestCase):
    """
    Test for cell_pattern_check function
    """

    def setUp(self) -> None:
        self.test_sheet = openpyxl.load_workbook(Path('.') / 'tests' / 'test' / 'test.xlsx').active
        self.correct_inputs = ('A1', 'AA999', 'ZZ1232', 'INPUT55555')
        self.incorrect_inputs = ['1A', 'ASD', '123', 'a1', '1A1']
        self.bad_inputs = [1, 0.5, [], {}, (), None, [None]]
        self.test_function = Functions('add', self.test_sheet)

    def test_cell_pattern_check_correct_inputs(self):
        expected = 0
        actual = self.test_function.cell_pattern_check(self.correct_inputs[0], self.correct_inputs[1],
                                                  self.correct_inputs[2], self.correct_inputs[3])
        self.assertEqual(expected, actual)

    def test_cell_pattern_check_incorrect_inputs(self):
        expected = 1
        for incorrect_input in self.incorrect_inputs:
            actual = self.test_function.cell_pattern_check(incorrect_input)
            self.assertEqual(expected, actual)

    def test_cell_pattern_check_correct_incorrect_inputs(self):
        expected = 1
        actual = self.test_function.cell_pattern_check(self.correct_inputs[0],
                                                  self.correct_inputs[1],
                                                  self.incorrect_inputs[2],
                                                  self.correct_inputs[3])
        self.assertEqual(expected, actual)

    def test_cell_pattern_check_bad_inputs(self):
        expected = -1
        for bad_input in self.bad_inputs:
            actual = self.test_function.cell_pattern_check(bad_input)
            self.assertEqual(expected, actual)


class CountingFunctionsTests(unittest.TestCase):
    """
    Tests for counting functions class
    """

    def setUp(self) -> None:
        self.test_sheet = openpyxl.load_workbook(Path('.') / 'tests' / 'test' / 'test.xlsx').active
        self.function_names = ['add', 'sub', 'multi', 'div', 'mean', 'incorrect']
        self.bad_inputs = ['string', 1, 0.5, [], {}, (), None, [None]]
        self.test_variables = ['A1', 'C1', 'E1', 'B2', '', '', '']
        self.test_function = CountingFunctions('add', self.test_sheet)
        self.test_function.variables = self.test_variables[0:3]
        self.test_function.get_attributes()

    def test_counting_functions_call(self):
        """
        Correct functions are executed
        """
        expected = 0
        for function_name in self.function_names[0:-1]:
            test_function = CountingFunctions(function_name, self.test_sheet)
            test_function.left_cell = self.test_variables[0]
            test_function.right_cell = self.test_variables[1]
            test_function.result_cell = self.test_variables[2]
            actual = test_function.function_call()
            self.assertEqual(expected, actual)

    def test_counting_functions_call_incorrect_function(self):
        """
        Test function call can handle incorrect functions names
        """
        expected = -1
        test_function = CountingFunctions(self.function_names[-1], self.test_sheet)
        actual = test_function.function_call()
        self.assertEqual(expected, actual)

    def test_counting_functions_call_bad_inputs(self):
        """
        Test function call can handle bad inputs
        """
        expected = -1
        for bad_input in self.bad_inputs:
            test_function = CountingFunctions(bad_input, self.test_sheet)
            test_function.left_cell = self.test_variables[0]
            test_function.right_cell = self.test_variables[1]
            test_function.result_cell = self.test_variables[2]
            actual = test_function.function_call()
            self.assertEqual(expected, actual)

    def test_get_attributes(self):
        """
        Test class ability to convert correct variables to attributes
        """
        expected = 0
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_wrong_quantity(self):
        """
        Test class can handle inconsistent number of attributes
        """
        expected = -1
        self.test_function.variables = self.test_variables
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_no_variables(self):
        """
        Test class can handle absence of variables
        """
        expected = -1
        self.test_function.variables = None
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_bad_inputs(self):
        """
        Test class can handle wrong variables' types
        """
        expected = -1
        self.test_function.variables = self.bad_inputs[0:3]
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_empty_coordinates(self):
        """
        Test class can handle empty inputs
        """
        expected = -1
        self.test_function.variables = self.test_variables[-3:]
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_add_ideal(self):
        """
        Ideal sum scenario
        """
        expected = 6
        actual = self.test_function.addition(self.test_function.get_worksheet(),
                                             self.test_function.left_cell,
                                             self.test_function.right_cell,
                                             self.test_function.result_cell)
        self.assertEqual(expected, actual)

    def test_addition_wrong_type(self):
        """
        No integers or floats.
        """
        expected = 0
        actual = self.test_function.addition(self.test_function.get_worksheet(), 'A6', 'C8', 'E1')
        self.assertEqual(expected, actual)

    def test_subtraction_ideal(self):
        """
        Ideal subtraction.
        """
        expected = -2
        actual = self.test_function.subtraction(self.test_function.get_worksheet(),
                                                'A1', 'C1', 'E1')
        self.assertEqual(expected, actual)

    def test_subtraction_wrong_type(self):
        """
        No integers or floats.
        """
        expected = 'N/A'
        actual = self.test_function.subtraction(self.test_function.get_worksheet(),
                                                'A1', 'C8', 'E1')
        self.assertEqual(expected, actual)

    def test_multiplication_ideal(self):
        """
        Ideal sum scenario. Simple.
        """
        expected = 6
        actual = self.test_function.multiplication(self.test_function.get_worksheet(),
                                                   'A1', 'C1', 'E1')
        self.assertEqual(expected, actual)

    def test_multiplication_wrong_type(self):
        """
        No integers or floats.
        """
        expected = 1
        actual = self.test_function.multiplication(self.test_function.get_worksheet(),
                                                   'A6', 'C8', 'E1')
        self.assertEqual(expected, actual)

    def test_division_ideal(self):
        """
        Ideal division.
        """
        expected = 3
        actual = self.test_function.division(self.test_function.get_worksheet(), 'C1', 'A1', 'E1')
        self.assertEqual(expected, actual)

    def test_division_wrong_type(self):
        """
        No integers or floats.
        """
        expected = 'N/A'
        actual = self.test_function.division(self.test_function.get_worksheet(), 'A6', 'C8', 'E1')
        self.assertEqual(expected, actual)

    def test_mean_ideal(self):
        """
        Ideal mean.
        """
        expected = 2
        actual = self.test_function.mean(self.test_function.get_worksheet(), 'A1', 'C1', 'E1')
        self.assertEqual(expected, actual)

    def test_mean_wrong_type(self):
        """
        No integers or floats.
        """
        expected = 'N/A'
        actual = self.test_function.mean(self.test_function.get_worksheet(), 'A6', 'C8', 'E1')
        self.assertEqual(expected, actual)


class DifficultFunctionsTests(unittest.TestCase):
    """
    Tests for difficult functions class
     """

    def setUp(self) -> None:
        self.test_sheet = openpyxl.load_workbook(Path('.') / 'tests' / 'test' / 'test.xlsx').active
        self.function_names = ['rounded', 'expo', 'log', 'incorrect']
        self.bad_inputs = ['string', [], {}, (), None, [None]]
        self.test_variables = ['A1', 2, 'E1', 'B2', '', '', '']
        self.test_function = DifficultFunctions('rounded', self.test_sheet)
        self.test_function.variables = self.test_variables[0:3]
        self.test_function.get_attributes()

    def test_difficult_functions_call(self):
        """
        Correct functions are executed
        """
        expected = 0
        for function_name in self.function_names[0:-1]:
            test_function = DifficultFunctions(function_name, self.test_sheet)
            test_function.cell = self.test_variables[0]
            test_function.number = self.test_variables[1]
            test_function.result_cell = self.test_variables[2]
            actual = test_function.function_call()
            self.assertEqual(expected, actual)

    def test_difficult_functions_call_incorrect_function(self):
        """
        Test function call can handle incorrect functions names
        """
        expected = -1
        test_function = DifficultFunctions(self.function_names[-1], self.test_sheet)
        actual = test_function.function_call()
        self.assertEqual(expected, actual)

    def test_difficult_functions_call_bad_inputs(self):
        """
        Test function call can handle bad inputs
        """
        expected = -1
        for bad_input in self.bad_inputs:
            test_function = DifficultFunctions(bad_input, self.test_sheet)
            test_function.cell = self.test_variables[0]
            test_function.number = self.test_variables[1]
            test_function.result_cell = self.test_variables[2]
            actual = test_function.function_call()
            self.assertEqual(expected, actual)

    def test_get_attributes(self):
        """
        Test class ability to convert correct variables to attributes
        """
        expected = 0
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_wrong_quantity(self):
        """
        Test class can handle inconsistent number of attributes
        """
        expected = -1
        self.test_function.variables = self.test_variables
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_no_variables(self):
        """
        Test class can handle absence of variables
        """
        expected = -1
        self.test_function.variables = None
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_bad_inputs(self):
        """
        Test class can handle wrong variables' types
        """
        expected = -1
        self.test_function.variables = self.bad_inputs[0:3]
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_empty_coordinates(self):
        """
        Test class can handle empty inputs
        """
        expected = -1
        self.test_function.variables = self.test_variables[-3:]
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_round_ideal(self):
        """
        Ideal round scenario
        """
        expected = 10.334
        actual = self.test_function.round_number(self.test_function.get_worksheet(), 'A4', 3, 'E1')
        self.assertEqual(expected, actual)

    def test_round_wrong_input(self):
        """
        Decimal is not int
        """
        expected = 10.33356
        actual = self.test_function.round_number(self.test_function.get_worksheet(), 'A4', [], 'E1')
        self.assertEqual(expected, actual)

    def test_exponent_ideal(self):
        """
        Ideal power scenario.
        """
        expected = 9
        actual = self.test_function.exponentiation(self.test_function.get_worksheet(),
                                                   'C1', 2, 'E1')
        self.assertEqual(expected, actual)

    def test_exponent_wrong_input(self):
        """
        Extension is not int
        """
        expected = 'N/A'
        actual = self.test_function.exponentiation(self.test_function.get_worksheet(),
                                                   'C1', [2], 'E1')
        self.assertEqual(expected, actual)

    def test_log_ideal(self):
        """
        Ideal logarithm scenario
        """
        expected = 1.0
        actual = self.test_function.logarithm(self.test_function.get_worksheet(),
                                              'C1', 3, 'E1')
        self.assertEqual(expected, actual)

    def test_log_wrong_input(self):
        """
        Wrong input scenario
        """
        expected = 'N/A'
        actual = self.test_function.logarithm(self.test_function.get_worksheet(),
                                              'C10', [3], 'E1')
        self.assertEqual(expected, actual)

    def test_log_wrong_base_input(self):
        """
        Wrong input scenario
        """
        expected = 'N/A'
        actual = self.test_function.logarithm(self.test_function.get_worksheet(),
                                              'C1', 1, 'E1')
        self.assertEqual(expected, actual)


class MovingFunctionsTests(unittest.TestCase):
    """
    Tests for moving functions class
    """

    def setUp(self) -> None:
        self.test_sheet = openpyxl.load_workbook(Path('.') / 'tests' / 'test' / 'test.xlsx').active
        self.function_names = ['move', 'copy', 'incorrect']
        self.bad_inputs = ['string', [], {}, (), None, [None]]
        self.test_variables = ['A1', 'C1', 'E1', 'B2', '', '', '']
        self.test_function = MovingFunctions('move', self.test_sheet)
        self.test_function.variables = self.test_variables[0:3]
        self.test_function.get_attributes()

    def test_moving_functions_call(self):
        """
        Correct functions are executed
        """
        expected = 0
        for function_name in self.function_names[0:-1]:
            test_function = MovingFunctions(function_name, self.test_sheet)
            test_function.left_corner = self.test_variables[0]
            test_function.right_corner = self.test_variables[1]
            test_function.result_cell = self.test_variables[2]
            actual = test_function.function_call()
            self.assertEqual(expected, actual)

    def test_difficult_functions_call_incorrect_function(self):
        """
        Test function call can handle incorrect functions names
        """
        expected = -1
        test_function = MovingFunctions(self.function_names[-1], self.test_sheet)
        actual = test_function.function_call()
        self.assertEqual(expected, actual)

    def test_difficult_functions_call_bad_inputs(self):
        """
        Test function call can handle bad inputs
        """
        expected = -1
        for bad_input in self.bad_inputs:
            test_function = MovingFunctions(bad_input, self.test_sheet)
            test_function.cell = self.test_variables[0]
            test_function.number = self.test_variables[1]
            test_function.result_cell = self.test_variables[2]
            actual = test_function.function_call()
            self.assertEqual(expected, actual)

    def test_get_attributes(self):
        """
        Test class ability to convert correct variables to attributes
        """
        expected = 0
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_wrong_quantity(self):
        """
        Test class can handle inconsistent number of attributes
        """
        expected = -1
        self.test_function.variables = self.test_variables
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_no_variables(self):
        """
        Test class can handle absence of variables
        """
        expected = -1
        self.test_function.variables = None
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_bad_inputs(self):
        """
        Test class can handle wrong variables' types
        """
        expected = -1
        self.test_function.variables = self.bad_inputs[0:3]
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_empty_coordinates(self):
        """
        Test class can handle empty inputs
        """
        expected = -1
        self.test_function.variables = self.test_variables[-3:]
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_move_ideal(self):
        """
        Ideal moving
        """
        expected = 'successful'
        actual = self.test_function.move(self.test_function.get_worksheet(), 'A1', 'C3', 'E1')
        self.assertEqual(expected, actual)

    def test_copy_ideal(self):
        """
        Ideal coping.
        """
        expected = 'successful'
        actual = self.test_function.copy(self.test_function.get_worksheet(), 'A1', 'C3', 'E1')
        self.assertEqual(expected, actual)


class DeleteFunctionTests(unittest.TestCase):
    """
    Tests for deletion function class
    """

    def setUp(self) -> None:
        self.test_sheet = openpyxl.load_workbook(Path('.') / 'tests' / 'test' / 'test.xlsx').active
        self.function_names = ['delete', 'incorrect']
        self.bad_inputs = ['string', [], {}, (), None, [None]]
        self.test_variables = ['A1', 'C1', '', '', '']
        self.test_function = DeleteFunction('delete', self.test_sheet)
        self.test_function.variables = self.test_variables[0:2]
        self.test_function.get_attributes()

    def test_delete_function_call(self):
        """
        Function is executed
        """
        expected = 0
        actual = self.test_function.function_call()
        self.assertEqual(expected, actual)

    def test_delete_function_call_incorrect_function(self):
        """
        Test function call can handle incorrect functions names
        """
        expected = -1
        test_function = DeleteFunction(self.function_names[-1], self.test_sheet)
        actual = test_function.function_call()
        self.assertEqual(expected, actual)

    def test_delete_function_call_bad_inputs(self):
        """
        Test function call can handle bad inputs
        """
        expected = -1
        for bad_input in self.bad_inputs:
            test_function = DeleteFunction(bad_input, self.test_sheet)
            test_function.left_corner = self.test_variables[0]
            test_function.right_corner = self.test_variables[1]
            actual = test_function.function_call()
            self.assertEqual(expected, actual)

    def test_get_attributes(self):
        """
        Test class ability to convert correct variables to attributes
        """
        expected = 0
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_wrong_quantity(self):
        """
        Test class can handle inconsistent number of attributes
        """
        expected = -1
        self.test_function.variables = self.test_variables
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_no_variables(self):
        """
        Test class can handle absence of variables
        """
        expected = -1
        self.test_function.variables = None
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_bad_inputs(self):
        """
        Test class can handle wrong variables' types
        """
        expected = -1
        self.test_function.variables = self.bad_inputs[0:2]
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_empty_coordinates(self):
        """
        Test class can handle empty inputs
        """
        expected = -1
        self.test_function.variables = self.test_variables[-2:]
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_delete_ideal(self):
        """
        Ideal scenario
        """
        expected = 'successful'
        actual = self.test_function.delete(self.test_function.get_worksheet(), 'A1', 'C3')
        self.assertEqual(expected, actual)


class CompareFunctionTests(unittest.TestCase):
    """
    Tests for deletion function class
    """

    def setUp(self) -> None:
        self.test_sheet = openpyxl.load_workbook(Path('.') / 'tests' / 'test' / 'test.xlsx').active
        self.function_names = ['compare', 'incorrect']
        self.bad_inputs = ['string', [], {}, (), None, [None]]
        self.test_variables = ['A1', 'C1', 'E1' '', '', '']
        self.test_function = CompareFunction('compare', self.test_sheet)
        self.test_function.variables = self.test_variables[0:3]
        self.test_function.get_attributes()

    def test_compare_function_call(self):
        """
        Function is executed
        """
        expected = 0
        actual = self.test_function.function_call()
        self.assertEqual(expected, actual)

    def test_compare_function_call_incorrect_function(self):
        """
        Test function call can handle incorrect functions names
        """
        expected = -1
        test_function = CompareFunction(self.function_names[-1], self.test_sheet)
        actual = test_function.function_call()
        self.assertEqual(expected, actual)

    def test_compare_function_call_bad_inputs(self):
        """
        Test function call can handle bad inputs
        """
        expected = -1
        for bad_input in self.bad_inputs:
            test_function = CompareFunction(bad_input, self.test_sheet)
            test_function.cell_1 = self.test_variables[0]
            test_function.cell_2 = self.test_variables[1]
            test_function.result_cell = self.test_variables[2]
            actual = test_function.function_call()
            self.assertEqual(expected, actual)

    def test_get_attributes(self):
        """
        Test class ability to convert correct variables to attributes
        """
        expected = 0
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_wrong_quantity(self):
        """
        Test class can handle inconsistent number of attributes
        """
        expected = -1
        self.test_function.variables = self.test_variables
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_no_variables(self):
        """
        Test class can handle absence of variables
        """
        expected = -1
        self.test_function.variables = None
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_bad_inputs(self):
        """
        Test class can handle wrong variables' types
        """
        expected = -1
        self.test_function.variables = self.bad_inputs[0:3]
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_empty_coordinates(self):
        """
        Test class can handle empty inputs
        """
        expected = -1
        self.test_function.variables = self.test_variables[-3:]
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_compare_ideal(self):
        """
        Ideal coping.
        """
        expected = True
        actual = self.test_function.compare(self.test_function.get_worksheet(), 'A1', 'A1', 'E1')
        self.assertEqual(expected, actual)

        expected = False
        actual = self.test_function.compare(self.test_function.get_worksheet(), 'A1', 'C1', 'E1')
        self.assertEqual(expected, actual)


class FindFunctionTests(unittest.TestCase):
    """
    Tests for deletion function class
    """

    def setUp(self) -> None:
        self.test_sheet = openpyxl.load_workbook(Path('.') / 'tests' / 'test' / 'test.xlsx').active
        self.function_names = ['find', 'incorrect']
        self.bad_inputs = ['string', [], {}, (), None, [None]]
        self.test_variables = ['find', '', '', '']
        self.test_function = SearchingFunction('find', self.test_sheet)
        self.test_function.variables = self.test_variables[0:1]
        self.test_function.get_attributes()

    def test_find_function_call(self):
        """
        Function is executed
        """
        expected = 0
        actual = self.test_function.function_call()
        self.assertEqual(expected, actual)

    def test_find_function_call_incorrect_function(self):
        """
        Test function call can handle incorrect functions names
        """
        expected = -1
        test_function = SearchingFunction(self.function_names[-1], self.test_sheet)
        actual = test_function.function_call()
        self.assertEqual(expected, actual)

    def test_find_function_call_bad_inputs(self):
        """
        Test function call can handle bad inputs
        """
        expected = -1
        for bad_input in self.bad_inputs:
            test_function = SearchingFunction(bad_input, self.test_sheet)
            test_function.element = self.test_variables[0]
            actual = test_function.function_call()
            self.assertEqual(expected, actual)

    def test_get_attributes(self):
        """
        Test class ability to convert correct variables to attributes
        """
        expected = 0
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_wrong_quantity(self):
        """
        Test class can handle inconsistent number of attributes
        """
        expected = -1
        self.test_function.variables = self.test_variables
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_no_variables(self):
        """
        Test class can handle absence of variables
        """
        expected = -1
        self.test_function.variables = None
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_get_attributes_bad_inputs(self):
        """
        Test class can handle wrong variables' types
        """
        expected = -1
        self.test_function.variables = self.bad_inputs[0:3]
        actual = self.test_function.get_attributes()
        self.assertEqual(expected, actual)

    def test_find_ideal(self):
        """
        Ideal find scenario
        """
        expected = ['B6', 'D6', 'F6']
        actual = self.test_function.find(self.test_function.get_worksheet(), 'find')
        self.assertEqual(expected, actual)

    def test_find_no_matches(self):
        """
        no matches find scenario
        """
        expected = 'no matches'
        actual = self.test_function.find(self.test_function.get_worksheet(), 'Z')
        self.assertEqual(expected, actual)
